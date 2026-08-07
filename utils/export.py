"""Shared export helpers: Excel (openpyxl) and PDF (fpdf2) report generation.

Used by pages/1_orders.py, pages/2_production.py, pages/3_procurement.py, and
pages/6_financials.py to let users download the table they're currently
looking at (filtered/visible rows), not the whole underlying table.
"""
import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Warm Chai theme amber accent (#C17F3E) as RGB
_AMBER = (193, 127, 62)
_INK = (44, 34, 24)      # #2C2218
_MUTED = (139, 106, 69)  # #8B6A45


def export_to_excel(df: pd.DataFrame, filename: str = "export.xlsx") -> io.BytesIO:
    """Render a DataFrame to a styled .xlsx file in memory."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = "Sheet1"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="C17F3E")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # .astype(str) leaves NaN/None as float NaN instead of stringifying it
            # (a pandas quirk) — fillna first so every value is a real string.
            values = df[col].fillna("").astype(str) if not df.empty else []
            max_len = max([len(str(col))] + [len(v) for v in values])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf


_CHAR_MAP = {
    "₹": "Rs. ",
    "—": "-", "–": "-",
    "…": "...",
    "✓": "v", "✕": "x", "•": "-",
}


def _pdf_safe(value) -> str:
    """fpdf2's core fonts (Helvetica) only support latin-1. Swap common
    Unicode punctuation used elsewhere in the app for ASCII equivalents,
    and fall back to '?' for anything else the font can't render."""
    if pd.isna(value):
        return "-"
    if isinstance(value, float) and value == int(value):
        value = int(value)
    text = str(value)
    for bad, good in _CHAR_MAP.items():
        text = text.replace(bad, good)
    try:
        text.encode("latin-1")
        return text
    except UnicodeEncodeError:
        return text.encode("latin-1", "replace").decode("latin-1")


def export_to_pdf(df: pd.DataFrame, title: str, filename: str = "export.pdf") -> io.BytesIO:
    """Render a DataFrame as a simple tabular PDF report."""
    n_cols = max(len(df.columns), 1)
    orientation = "L" if n_cols > 6 else "P"
    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_title(title)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*_AMBER)
    pdf.cell(0, 10, _pdf_safe(title), ln=1)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*_MUTED)
    pdf.cell(0, 6, f"Generated {datetime.now().strftime('%d %b %Y %H:%M')} · {len(df)} rows", ln=1)
    pdf.ln(3)

    page_width = pdf.w - pdf.l_margin - pdf.r_margin
    col_width = page_width / n_cols

    def _fit(text: str) -> str:
        """Truncate text with an ellipsis so it never overflows its cell width
        (fpdf2's cell() doesn't clip/wrap, so overflow silently overlaps the
        next cell and comes out garbled when a PDF viewer extracts the text)."""
        pad = 2  # mm of breathing room inside the cell border
        max_w = col_width - pad
        if pdf.get_string_width(text) <= max_w:
            return text
        while text and pdf.get_string_width(text + "...") > max_w:
            text = text[:-1]
        return text + "..." if text else "..."

    def _header_row():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*_AMBER)
        pdf.set_text_color(255, 255, 255)
        for col in df.columns:
            pdf.cell(col_width, 8, _fit(_pdf_safe(col)), border=1, fill=True, align="C")
        pdf.ln()

    _header_row()

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_INK)
    fill = False
    for _, row in df.iterrows():
        if pdf.get_y() > pdf.h - pdf.b_margin - 10:
            pdf.add_page()
            _header_row()
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(*_INK)
        pdf.set_fill_color(255, 248, 240)  # #FFF8F0 zebra stripe
        for val in row:
            pdf.cell(col_width, 7, _fit(_pdf_safe(val)), border=1, fill=fill, align="L")
        pdf.ln()
        fill = not fill

    buf = io.BytesIO(bytes(pdf.output()))
    buf.seek(0)
    return buf


_LOGO_PATH_D = "M 50 10 A 40 40 0 1 0 85 65"  # matches the app's SVG logo arc


def _draw_logo(pdf: FPDF, x: float, y: float, size: float = 16) -> None:
    """Approximate the app's circular-arrow SVG logo using fpdf2 vector primitives."""
    r = size / 2
    cx, cy = x + r, y + r
    pdf.set_draw_color(*_AMBER)
    pdf.set_line_width(size * 0.12)
    # Arc: ~300° open circle (leaves a gap, like the logo's "C" shape)
    pdf.ellipse(x, y, size, size, style="D")
    pdf.set_fill_color(*_AMBER)
    pdf.ellipse(cx - size * 0.09, cy - size * 0.09, size * 0.18, size * 0.18, style="F")


def generate_dispatch_note(order: dict, filename: str = "dispatch_note.pdf") -> io.BytesIO:
    """Generate a single-page operational dispatch slip for one order.

    Not a tax invoice — no GST fields, no invoice numbering. `order` should
    contain: id, company, customer_name, product, quantity, quantity_unit,
    rate, rate_type, order_date, expected_dispatch_date, actual_dispatch_date
    (optional), batch_reference (optional), and optionally company_address /
    company_gstin / company_contact (from Settings -> Company Details; shown
    as blank/omitted if not configured there).
    """
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_title(f"Dispatch Note - Order #{order.get('id', '')}")
    pdf.add_page()

    _draw_logo(pdf, pdf.l_margin, 12, size=16)
    pdf.set_xy(pdf.l_margin + 22, 12)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*_AMBER)
    pdf.cell(0, 10, "Reclaimr", ln=1)
    pdf.set_x(pdf.l_margin + 22)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*_MUTED)
    pdf.cell(0, 6, "Manufacturing ERP", ln=1)

    pdf.ln(12)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(*_INK)
    pdf.cell(0, 10, "Dispatch Note", ln=1)
    pdf.set_draw_color(232, 213, 183)  # #E8D5B7
    pdf.set_line_width(0.4)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(6)

    rate_label = "per unit" if order.get("rate_type") == "per_unit" else "overall total"
    company_line = order.get("company") or "—"
    company_extra = ", ".join(filter(None, [
        order.get("company_address"),
        f"GSTIN: {order.get('company_gstin')}" if order.get("company_gstin") else None,
        f"Contact: {order.get('company_contact')}" if order.get("company_contact") else None,
    ]))
    if company_extra:
        company_line = f"{company_line} ({company_extra})"

    fields = [
        ("Order ID", f"#{order.get('id', '—')}"),
        ("Company", company_line),
        ("Customer", order.get("customer_name") or "—"),
        ("Product", order.get("product") or "—"),
        ("Quantity", f"{order.get('quantity', '—')} {order.get('quantity_unit', '') or ''}".strip()),
        ("Rate", f"Rs. {order.get('rate', '—')} ({rate_label})"),
        ("Order Date", order.get("order_date") or "—"),
        ("Expected Dispatch", order.get("expected_dispatch_date") or "—"),
        ("Actual Dispatch", order.get("actual_dispatch_date") or "—"),
        ("Batch Reference", order.get("batch_reference") or "—"),
    ]

    for label, val in fields:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*_MUTED)
        pdf.cell(50, 9, f"{label}", border=0)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(*_INK)
        pdf.cell(0, 9, _pdf_safe(val), ln=1)

    pdf.ln(14)
    pdf.set_draw_color(232, 213, 183)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(*_MUTED)
    pdf.multi_cell(0, 6, "This is a dispatch confirmation, not a tax invoice.")

    buf = io.BytesIO(bytes(pdf.output()))
    buf.seek(0)
    return buf
